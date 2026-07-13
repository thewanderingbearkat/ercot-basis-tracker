"""Build Seller_Screen_v3.xlsx: Read Me (live weights) | Screen (ranked, curated,
live composite + source lookups) | All Plants | 6 source tabs | Owner Tiers."""
import os
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DESK = r"C:\Users\TylerMartin\OneDrive - ArcLight Renewable Services\Desktop"
SCRATCH = os.environ.get("MA_SCREEN_WORKDIR", r"C:\ma_screen_work")  # workdir with component CSVs
OUT = os.path.join(DESK, "Seller_Screen_v3.xlsx")

F = "Arial"
NAVY, LTGRAY, BLUE = "1F3864", "F2F2F2", "0000FF"
H_FILL = PatternFill("solid", start_color=NAVY)
G_FILL = PatternFill("solid", start_color="D9E2F3")
NOTE_FONT = Font(name=F, size=9, italic=True, color="595959")
HDR_FONT = Font(name=F, size=10, bold=True, color="FFFFFF")
THIN = Border(bottom=Side(style="thin", color="BFBFBF"))

with open(os.path.join(SCRATCH, "latest_scores_path.txt")) as fh:
    scores_path = fh.read().strip()
final = pd.read_csv(scores_path, encoding="utf-8-sig")
comp = {n: pd.read_csv(os.path.join(SCRATCH, f), encoding="utf-8-sig").drop_duplicates("POWER_PLANT")
        for n, f in [("eqr", "eqr_plant_labels.csv"), ("sellers", "seller_signals.csv"),
                     ("dc", "dc_proximity.csv"), ("bf", "bustedflip_port.csv"),
                     ("cong", "congestion_overlay.csv"), ("own", "flip_precision.csv"),
                     ("negp", "negprice_overlay.csv")]}
t50 = pd.read_excel(os.path.join(DESK, "Top 50 Asset Owners_7.1.2026.xlsx"),
                    sheet_name="Top 50 Owners", header=4)

wb = Workbook()

def style_header(ws, row, ncols, fill=H_FILL, font=HDR_FONT):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill, cell.font = fill, font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

DATA_FONT = Font(name=F, size=9)

def dump_df(ws, df, header_row, widths=None, max_col_width=42):
    df = df.replace({np.nan: None})
    for j, col in enumerate(df.columns, 1):
        ws.cell(row=header_row, column=j, value=str(col))
    for i, row in enumerate(df.itertuples(index=False), header_row + 1):
        for j, v in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = DATA_FONT
    style_header(ws, header_row, len(df.columns))
    for j, col in enumerate(df.columns, 1):
        w = min(max(len(str(col)) + 2, 12), max_col_width)
        ws.column_dimensions[get_column_letter(j)].width = (widths or {}).get(col, w)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(df.columns))}{header_row + len(df)}"
    ws.freeze_panes = f"A{header_row + 1}"

# ============ Read Me ============
rm = wb.active
rm.title = "Read Me"
rm.sheet_view.showGridLines = False
for col, w in zip("ABCDE", (2, 30, 14, 90, 8)):
    rm.column_dimensions[col].width = w

def put(ws, r, c, v, bold=False, size=10, color="000000", wrap=True, italic=False, fill=None,
        fmt=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name=F, size=size, bold=bold, color=color, italic=italic)
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    if fill:
        cell.fill = PatternFill("solid", start_color=fill)
    if fmt:
        cell.number_format = fmt
    return cell

put(rm, 2, 2, "SkyVest — Willing-Seller Screen v3", bold=True, size=16, color=NAVY)
put(rm, 3, 2, "Ranks operating renewables (>=50 MW) by likelihood the owner is a willing seller "
              "in a bilateral, undervalued / distressed situation. Composite is LIVE: edit the "
              "blue weights below and the Screen tab re-ranks.", italic=True)

put(rm, 5, 2, "WEIGHTS (blue cells — must sum to 1.00)", bold=True, size=11, color=NAVY)
WROWS = {"Contract posture": (6, "Merchant / PPA-expiring / EQR roll-off. 100=confirmed roll-off, 95=EQR merchant, 90=S&P no-PPA, 80=PPA exp <=3yr, 65=partially contracted, 60=exp <=5yr or verify-conflict, 40=not assessed, 15=EQR contracted, 10=long-dated PPA.", "Src EQR + All Plants (S&P PPA cols)"),
         "Busted-flip fingerprint": (7, "v2 Co-Own rubric (4+ owners=100, 3=85, 2=65, 1=15/45/5) lifted to >=90 when stake data shows sponsor <=10% beside fund >=60%, >=80 when co-owners entered on one date/KeyDeal (single financing close).", "Src BustedFlip + Src Ownership"),
         "Vintage / PTC": (8, "Wind: COD<=2016 & not repowered=100 (PTC fully expired), 2017-18=50, newer=10, repowered=15. Solar: <=2016=70, <=2019=55, <=2021=45 (ITC recapture cleared), newer=10. Battery=10.", "All Plants (COD, repower flags)"),
         "Willing-seller behavior": (9, "Live sale process=100 (owner) / 70 (parent-only); distressed=95/80; repeat seller=80/50; sold once=60/40; financial owner floor=55; asset traded 2024+ caps at 30 (just bought); traded <=2023 lifts to 65 (fund-recycle).", "Src Sellers"),
         "CF vs cohort P50": (10, "v2 rubric: CF'23-24 / cohort-P50 ratio. <=0.5=100, <=0.7=85, <=0.85=65, <=1.0=45, <=1.15=25, else 10. Too new / no data = neutral 40.", "Src BustedFlip"),
         "Opportunity tags": (11, "Sum of tag points (table below) for every tag the plant trips, capped at 100. Flags are LIVE formulas on the Tag Audit tab.", "Tag Audit")}
W = [0.21, 0.17, 0.17, 0.17, 0.13, 0.15]
for (name, (r, rub, src)), w in zip(WROWS.items(), W):
    put(rm, r, 2, name, bold=True)
    c = put(rm, r, 3, w, fmt="0.00")
    c.font = Font(name=F, size=10, color=BLUE, bold=True)
    c.fill = PatternFill("solid", start_color="DDEBF7")
    put(rm, r, 4, rub + "  [" + src + "]", size=9)
    rm.row_dimensions[r].height = 34
put(rm, 12, 2, "Sum (must = 1.00)", bold=True, italic=True)
sc = rm.cell(row=12, column=3, value="=SUM(C6:C11)")
sc.number_format = "0.00"
sc.font = Font(name=F, size=10, bold=True)
from openpyxl.formatting.rule import CellIsRule
rm.conditional_formatting.add("C12", CellIsRule(
    operator="notEqual", formula=["1"], fill=PatternFill("solid", start_color="F4CCCC")))

# TAG POINTS — blue editable cells C15:C24; order MUST match Tag Audit flag columns
TAGS = [("capitulation", 40, "merchant/PPA-expiring AND congestion worsening >$1/MWh y/y — both revenue legs deteriorating"),
        ("cf-collapse", 35, "2024 CF < 75% of 2023 (full-baseline, non-repowered plants only; 94% EIA-validated)"),
        ("mechanical-fix", 25, "CF underperformer at a benign node — availability problem, buy-and-fix thesis"),
        ("curtailment-play", 25, "CF underperformer at a congested node — economic curtailment, congestion-optionality/repower thesis"),
        ("neg-price-bleed", 30, "PTC-expired wind with >=15% negative-price DA hours — cash-negative hours w/ no PTC offset"),
        ("yieldco-stress", 30, "owned by a capital-stressed public vehicle (XPLR/Clearway/Atlantica...) — top-50 carve-out, motivated seller"),
        ("fund-life-exit", 25, "financial owner in the position since <=2018 — past typical fund term"),
        ("oem-orphan", 20, "turbines from a dead/exited OEM (Suzlon, Clipper, Senvion, Mitsubishi...) — O&M cost spiral, repower case"),
        ("busted-flip", 25, "flip pillar >=85 — tax-equity structure past its scheduled flip"),
        ("ptc-merchant", 30, "PTC expired AND merchant/PPA-expiring — sponsor economics dead")]
TAG_PT_ROW0 = 15  # first tag-points row on Read Me
put(rm, 14, 2, "TAG POINTS (blue — each tripped tag adds points to the Opportunity-tags "
               "pillar, capped at 100; edit to re-rank)", bold=True, size=11, color=NAVY)
for i, (tag, pts, desc) in enumerate(TAGS):
    r = TAG_PT_ROW0 + i
    put(rm, r, 2, tag, bold=True)
    c = put(rm, r, 3, pts, fmt="0")
    c.font = Font(name=F, size=10, color=BLUE, bold=True)
    c.fill = PatternFill("solid", start_color="DDEBF7")
    put(rm, r, 4, desc, size=9)

put(rm, 26, 2, "HOW TO AUDIT A SCORE", bold=True, size=11, color=NAVY)
put(rm, 27, 2, "1. Screen tab: every pillar has an evidence column beside it (plain-English reason).\n"
               "2. EQR label, congestion, neg-price and data-center cells on Screen are live VLOOKUPs "
               "into the source tabs — trace any value by finding the plant row on the named Src tab.\n"
               "3. Tag Audit tab is the tag backup: raw inputs on the left, one LIVE 0/1 formula per "
               "tag (open any flag cell to see its exact rule), and the points-weighted tag score on "
               "the right. Screen's 'Tags' pillar reads from it; editing tag points above re-scores "
               "everything.\n"
               "4. Src tabs are verbatim outputs of the scoring pipeline (assemble_v3.py), keyed by "
               "plant name. All Plants holds the full 2,221-plant universe incl. excluded rows.\n"
               "5. The five classic pillar sub-scores are pipeline-computed values (rubrics above); "
               "weights, tag flags, tag score and composite are live Excel formulas.")
rm.merge_cells("B27:D27")
rm.row_dimensions[27].height = 110

put(rm, 29, 2, "DATA SOURCES (all exported / pulled 2026-07)", bold=True, size=11, color=NAVY)
SRC = [("Src EQR", "FERC EQR via Snowflake AMA.EQR (2024+). Seller LLC -> buyer, ISO share, price. Blind in ERCOT (non-jurisdictional). 921 plants matched."),
       ("Src Sellers", "transaction.csv + projects-export.csv deal intel: live processes, repeat sellers, distress, asset trade history."),
       ("Src BustedFlip", "SkyVest_BustedFlip_Screen_v2.xlsx ported by name (1,289 plants): Co-Own / CF scores, cohort P50, TE-owner inference."),
       ("Src Ownership", "SkyVest_Ownership_Records_AllUS_ISO.xlsx stake data: financing-event dates, sponsor/TE structure, EIA codes. Wind+solar ISO plants only."),
       ("Src Congestion", "Yes Energy via Snowflake: plant->EIA->node->DA/RT congestion 2024-25 ($/MWh). ERCOT = node minus HB_HUBAVG basis. 1,405 plants priced."),
       ("Src DataCenters", "data-center.csv pipeline within 50/100 mi (haversine), capped 2 GW/project, UC x0.9 / planned x0.4, crypto x0.5."),
       ("Src NegPrice", "Yes Energy DART_PRICES: share of hours w/ negative DA/RT LMP at the plant's node, 2024/2025. Flag at >=15% of DA hours."),
       ("Owner Tiers", "Top 50 Asset Owners_7.1.2026.xlsx — exclusion list driving filter_pass."),
       ("S&P screens", "Asset Screen_7.1.2026.xlsx (universe, PPA fields) + Unit Screen (repower, equipment) — carried on All Plants.")]
for i, (tab, desc) in enumerate(SRC):
    put(rm, 30 + i, 2, tab, bold=True)
    put(rm, 30 + i, 4, desc, size=9)
    rm.row_dimensions[30 + i].height = 24

r0 = 30 + len(SRC) + 1
put(rm, r0, 2, "CAVEATS", bold=True, size=11, color=NAVY)
CAV = ["CF confidence: S&P CF was validated against raw EIA-923 (66-plant sample): ~90% agree within 3pp (median deviation 0.0pp) and 29/31 cf-collapse flags confirmed independently. Known artifact classes — S&P capacity vs EIA-site nameplate mismatches on multi-phase plants, repowered plants, and <2yr baselines — are excluded from the cf-collapse flag by construction.",
       "Joins are plant-name keyed except congestion (EIA-keyed). Expect a few % mismatch; EQR matches contradicting S&P PPA data are scored 60 with a VERIFY reason, not treated as merchant.",
       "EQR cannot see ERCOT contracts; ERCOT contract scores rest on S&P fields alone. Battery tolling is invisible to both.",
       "~1/3 of plants carry neutral-40 CF (too new / unscored by v2). Congestion is a time-average (not generation-weighted); ERCOT basis includes the small loss component.",
       "Ownership export holds only CURRENT stakes — co-owner-exit signal is empty until re-exported with ended stakes. No batteries / non-ISO plants in that export.",
       "Data-center pipeline MW is announcement-inflated; treat dc_score 3 + single mega-project with suspicion (see per-project cap note on Src DataCenters)."]
for i, c in enumerate(CAV):
    put(rm, r0 + 1 + i, 2, f"{i + 1}.", bold=True)
    put(rm, r0 + 1 + i, 4, c, size=9)
    rm.row_dimensions[r0 + 1 + i].height = 24 if len(c) < 200 else 36

# ============ Tag Audit ============
TA_INPUTS = [  # header, final-df column
    ("Plant", "plant"), ("Contract score", "score_contract"),
    ("Cong trend $/MWh", "cong_trend"), ("CF score", "score_cf"),
    ("DA cong '25", "cong_da_25"), ("CF 2023 %", "cf_2023"), ("CF 2024 %", "cf_2024"),
    ("COD", "cod_yr"), ("Repowered", "repowered"), ("PTC expired", "ptc_expired"),
    ("Neg px share '25", "negp_da_25"), ("Stressed top-50", "top50_stressed"),
    ("Financial owner", "financial_owner"), ("Oldest stake yr", "oldest_current_stake_yr"),
    ("OEM orphan", "oem_orphan"), ("OEM makes", "oem_mfrs"), ("Flip score", "score_flip"),
]
# per-tag flag formulas over the input columns above (same rules as assemble_v3.py / Read Me)
TA_FLAGS = [
    ("capitulation", "=IF(AND($B{x}>=80,ISNUMBER($C{x}),$C{x}<-1),1,0)"),
    ("cf-collapse", "=IF(AND(ISNUMBER($F{x}),ISNUMBER($G{x}),$F{x}>5,$G{x}<0.75*$F{x},"
                    "ISNUMBER($H{x}),$H{x}<=2022,$I{x}=0),1,0)"),
    ("mechanical-fix", "=IF(AND($D{x}>=65,ISNUMBER($E{x}),$E{x}>-3),1,0)"),
    ("curtailment-play", "=IF(AND($D{x}>=65,ISNUMBER($E{x}),$E{x}<=-3),1,0)"),
    ("neg-price-bleed", "=IF(AND($J{x}=1,ISNUMBER($K{x}),$K{x}>=0.15),1,0)"),
    ("yieldco-stress", "=IF($L{x}=1,1,0)"),
    ("fund-life-exit", "=IF(AND($M{x}=1,ISNUMBER($N{x}),$N{x}<=2018),1,0)"),
    ("oem-orphan", "=IF($O{x}=1,1,0)"),
    ("busted-flip", "=IF($Q{x}>=85,1,0)"),
    ("ptc-merchant", "=IF(AND($J{x}=1,$B{x}>=80),1,0)"),
]
assert [t for t, _ in TA_FLAGS] == [t for t, _, _ in TAGS]  # row/col order must match Read Me
ta = wb.create_sheet("Tag Audit")
put(ta, 1, 1, "Tag backup: inputs (left, static pipeline values) -> one LIVE 0/1 formula per tag "
              "(open a cell to see the exact rule) -> tag score = sum of tripped tags' points "
              "('Read Me' C15:C24), capped at 100. Screen's 'Tags' pillar reads this tab.",
    italic=True, size=9, color="595959")
nfix = len(TA_INPUTS)
ta_headers = [h for h, _ in TA_INPUTS] + [f"tag: {t}" for t, _ in TA_FLAGS] + ["Tag score (0-100)"]
for j, h in enumerate(ta_headers, 1):
    ta.cell(row=2, column=j, value=h)
    ta.column_dimensions[get_column_letter(j)].width = 13 if j > 1 else 38
style_header(ta, 2, len(ta_headers))
ta_recs = final.replace({np.nan: None}).to_dict("records")
score_terms = " + ".join(
    f"{get_column_letter(nfix + 1 + i)}{{x}}*'Read Me'!$C${TAG_PT_ROW0 + i}"
    for i in range(len(TA_FLAGS)))
for i, r in enumerate(ta_recs):
    x = i + 3
    for j, (_, col) in enumerate(TA_INPUTS, 1):
        c = ta.cell(row=x, column=j, value=r.get(col))
        c.font = DATA_FONT
    for k, (_, fml) in enumerate(TA_FLAGS):
        c = ta.cell(row=x, column=nfix + 1 + k, value=fml.format(x=x))
        c.font = DATA_FONT
        c.number_format = "0"
    c = ta.cell(row=x, column=len(ta_headers), value=f"=MIN(100,{score_terms.format(x=x)})")
    c.font = Font(name=F, size=9, bold=True)
    c.number_format = "0"
ta.freeze_panes = "B3"
ta.auto_filter.ref = f"A2:{get_column_letter(len(ta_headers))}{len(ta_recs) + 2}"
TA_SCORE_IX = len(ta_headers)
TA_LAST = get_column_letter(TA_SCORE_IX)

# ============ Screen ============
sc_df = final[final.filter_pass == 1].sort_values("composite", ascending=False).reset_index(drop=True)
ws = wb.create_sheet("Screen")
ws.sheet_view.showGridLines = False

def col_of(df, name):
    return df.columns.get_loc(name) + 1

eqr_ix = col_of(comp["eqr"], "eqr_label")
cong_ix = {k: col_of(comp["cong"], k) for k in ("cong_da_25", "cong_trend")}
dc_ix = col_of(comp["dc"], "dc_mw_50")

def glook(tab, df, ix):
    """Blank-safe lookup: blank source cells show em-dash, not a fabricated 0."""
    last = get_column_letter(len(df.columns))
    def fn(r, x, v=None):
        lk = f"VLOOKUP($C{x},'{tab}'!$A:${last},{ix},FALSE)"
        return f'=IFERROR(IF({lk}="","—",{lk}),"—")'
    return fn

SPEC = [  # header, width, fmt, value fn (row dict, excel row) -> value
    ("Rank", 6, "0", lambda r, x: f"=RANK($B{x},$B$3:$B${len(sc_df) + 2})"),
    ("Composite", 10, "0.0", lambda r, x: f"='Read Me'!$C$6*$I{x}+'Read Me'!$C$7*$J{x}"
                                          f"+'Read Me'!$C$8*$K{x}+'Read Me'!$C$9*$L{x}"
                                          f"+'Read Me'!$C$10*$M{x}+'Read Me'!$C$11*$N{x}"),
    ("Plant", 38, None, lambda r, x: r["plant"]),
    ("Tech", 8, None, lambda r, x: r["tech"]),
    ("MW", 8, "#,##0", lambda r, x: r["capacity_mw"]),
    ("ISO", 10, None, lambda r, x: r["iso"]),
    ("St", 5, None, lambda r, x: r["state"]),
    ("COD", 7, "0", lambda r, x: r["cod_yr"]),
    ("Contract", 9, "0", lambda r, x: r["score_contract"]),
    ("Flip", 6, "0", lambda r, x: r["score_flip"]),
    ("Vintage", 8, "0", lambda r, x: r["score_vintage"]),
    ("Seller", 7, "0", lambda r, x: r["score_seller"]),
    ("CF", 5, "0", lambda r, x: r["score_cf"]),
    ("Tags", 6, "0", lambda r, x: f"=IFERROR(VLOOKUP($C{x},'Tag Audit'!$A:${TA_LAST},"
                                  f"{TA_SCORE_IX},FALSE),0)"),
    ("Contract evidence", 46, None, lambda r, x: r["contract_reason"]),
    ("PPA expiry", 11, "yyyy-mm-dd",
     lambda r, x: pd.to_datetime(r["ppa_expiry"]).date() if r["ppa_expiry"] else None),
    ("EQR label", 12, None, lambda r, x: f"=IFERROR(VLOOKUP($C{x},'Src EQR'!$A:$"
        f"{get_column_letter(len(comp['eqr'].columns))},{eqr_ix},FALSE),\"—\")"),
    ("Flip / TE evidence", 46, None,
     lambda r, x: (r["flip_reason"] if isinstance(r["flip_reason"], str) and r["flip_reason"]
                   else (r["bf_te_owner"] if isinstance(r["bf_te_owner"], str) else ""))),
    ("Seller evidence", 46, None, lambda r, x: r["seller_reason"]),
    ("Opportunity tags", 34, None, lambda r, x: r["opportunity_tags"]),
    ("PTC done", 9, None, lambda r, x: "Y" if r["ptc_expired"] == 1 else "—"),
    ("Repow", 7, None, lambda r, x: "Y" if r["repowered"] == 1 else "—"),
    ("CF '23-24", 9, "0.0", lambda r, x: r["bf_cf_2324"]),
    ("P50", 6, "0.0", lambda r, x: r["bf_cohort_p50"]),
    ("Cong 0-3", 9, "0", lambda r, x: r["cong_score"]),
    ("DA cong '25 $/MWh", 11, "0.00;[Red]-0.00",
     glook("Src Congestion", comp["cong"], cong_ix["cong_da_25"])),
    ("Cong trend $/MWh (+=better)", 13, "0.00;[Red]-0.00",
     glook("Src Congestion", comp["cong"], cong_ix["cong_trend"])),
    ("Neg px hrs '25", 9, "0%",
     glook("Src NegPrice", comp["negp"], col_of(comp["negp"], "negp_da_25"))),
    ("DC 0-3", 6, "0", lambda r, x: r["dc_score"]),
    ("DC MW 50mi", 11, "#,##0", glook("Src DataCenters", comp["dc"], dc_ix)),
    ("Owners", 50, None, lambda r, x: r["owners"]),
    ("Ultimate parents", 45, None, lambda r, x: r["ult_parents"]),
]
GROUPS = [("RANKING", 1, 2), ("ASSET", 3, 8), ("PILLAR SCORES (0-100)", 9, 14),
          ("EVIDENCE", 15, 24), ("OVERLAYS (not weighted)", 25, 30), ("OWNERSHIP", 31, 32)]
for title, c1, c2 in GROUPS:
    ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
    cell = ws.cell(row=1, column=c1, value=title)
    cell.font = Font(name=F, size=9, bold=True, color=NAVY)
    cell.fill = G_FILL
    cell.alignment = Alignment(horizontal="center")
for j, (h, wdt, _, _) in enumerate(SPEC, 1):
    ws.cell(row=2, column=j, value=h)
    ws.column_dimensions[get_column_letter(j)].width = wdt
style_header(ws, 2, len(SPEC))
PILLAR_NOTES = {"Contract": "0.25 wt — see Read Me rubric; evidence in 'Contract evidence' + Src EQR",
                "Flip": "0.20 wt — v2 Co-Own + stake fingerprints; see 'Flip / TE evidence' + Src Ownership",
                "Vintage": "0.20 wt — PTC/ITC rubric on Read Me; COD + Repow cols",
                "Seller": "0.20 wt — deal-intel signals; see 'Seller evidence' + Src Sellers",
                "CF": "0.13 wt — CF'23-24 vs cohort P50; Src BustedFlip",
                "Tags": "0.15 wt — sum of tripped tags' points (Read Me C15:C24), cap 100; live formula chain on Tag Audit tab"}
for j, (h, *_ ) in enumerate(SPEC, 1):
    if h in PILLAR_NOTES:  # exact match only — evidence columns must not inherit pillar notes
        ws.cell(row=2, column=j).comment = Comment(PILLAR_NOTES[h], "Screen v3", height=90, width=260)

recs = sc_df.replace({np.nan: None}).to_dict("records")
for i, r in enumerate(recs):
    x = i + 3
    for j, (_, _, fmt, fn) in enumerate(SPEC, 1):
        v = fn(r, x)
        cell = ws.cell(row=x, column=j, value=v)
        cell.font = Font(name=F, size=9)
        cell.border = THIN
        cell.alignment = Alignment(vertical="top", wrap_text=SPEC[j - 1][1] >= 38)
        if fmt:
            cell.number_format = fmt
last = len(recs) + 2
ws.auto_filter.ref = f"A2:{get_column_letter(len(SPEC))}{last}"
ws.freeze_panes = "D3"
def hdr_col(name):
    return get_column_letter([h for h, *_ in SPEC].index(name) + 1)

ws.conditional_formatting.add(f"B3:B{last}", ColorScaleRule(
    start_type="min", start_color="FFFFFF", end_type="max", end_color="70AD47"))
ws.conditional_formatting.add(f"I3:N{last}", ColorScaleRule(
    start_type="num", start_value=0, start_color="FFFFFF",
    end_type="num", end_value=100, end_color="9DC3E6"))
cg = hdr_col("Cong 0-3")
ws.conditional_formatting.add(f"{cg}3:{cg}{last}", ColorScaleRule(
    start_type="num", start_value=0, start_color="FFFFFF",
    end_type="num", end_value=3, end_color="E06666"))
dcc = hdr_col("DC 0-3")
ws.conditional_formatting.add(f"{dcc}3:{dcc}{last}", ColorScaleRule(
    start_type="num", start_value=0, start_color="FFFFFF",
    end_type="num", end_value=3, end_color="FFD966"))

# ============ All Plants ============
ap = wb.create_sheet("All Plants")
put(ap, 1, 1, "Full >=50 MW universe (2,221 plants) — includes excluded rows: filter_pass=0 = "
              "top-50-owned or outside footprint ISOs. Static snapshot from assemble_v3.py: "
              "composite/rank here do NOT respond to weight edits (only the Screen tab is live).",
    italic=True, size=9, color="595959")
dump_df(ap, final, 2)

# ============ Source tabs ============
SRC_TABS = [("Src EQR", "eqr", "FERC EQR seller classification, 2024+ (Snowflake AMA.EQR). MERCHANT = >=80% of MWh to ISO; CONTRACTED = bilateral offtake; eqr_rolloff = bilateral sales ended, now spot."),
            ("Src Sellers", "sellers", "Owner willing-seller signals from transaction.csv / projects-export.csv. seller_owner_matched shows which owner/parent tripped the signal."),
            ("Src BustedFlip", "bf", "Ported from SkyVest_BustedFlip_Screen_v2.xlsx Target Screen (exact name matches only)."),
            ("Src Ownership", "own", "Stake-level fingerprints from SkyVest_Ownership_Records_AllUS_ISO.xlsx. flip_event = co-owners share one Own Event Date / KeyDeal. sponsor_small_stake = <=10% beside >=60%."),
            ("Src Congestion", "cong", "Yes Energy DART_PRICES 2024/2025 averages. Non-ERCOT = DA/RT congestion component; ERCOT = node minus HB_HUBAVG basis (incl. losses). Negative = costs the generator."),
            ("Src DataCenters", "dc", "Data-center pipeline proximity. dc_mw_50 = weighted MW (UC x0.9, planned x0.4, crypto x0.5, 2 GW/project cap, accuracy A/B) within 50 mi."),
            ("Src NegPrice", "negp", "Share of hours with negative DA/RT LMP at the plant's node (Yes Energy DART_PRICES, 2024/2025). negp_hours_flag = 1 when >=15% of 2025 DA hours were negative.")]
for tab, key, note in SRC_TABS:
    s = wb.create_sheet(tab)
    put(s, 1, 1, note, italic=True, size=9, color="595959")
    dump_df(s, comp[key], 2)

# ============ Owner Tiers ============
ot = wb.create_sheet("Owner Tiers")
put(ot, 1, 1, "Top-50 owners (by net MW) — plants whose owner or ultimate parent matches this list "
              "get filter_pass=0 on All Plants / are excluded from Screen. Matching uses "
              "normalized tokens + alias list (see assemble_v3.py).", italic=True, size=9,
    color="595959")
rank_col = t50.columns[0]
t50_data = t50[t50[rank_col].notna()]
t50_tot = t50[t50[rank_col].isna()]
dump_df(ot, t50_data, 2)  # total row kept OUT of the filter range so sorting can't mix it in
tr = 2 + len(t50_data) + 1
for j, v in enumerate(t50_tot.replace({np.nan: None}).itertuples(index=False).__next__()
                      if len(t50_tot) else [], 1):
    c = ot.cell(row=tr, column=j, value=v)
    c.font = Font(name=F, size=9, bold=True)
    c.border = Border(top=Side(style="medium", color=NAVY))

wb.move_sheet("Tag Audit", offset=1)  # Read Me | Screen | Tag Audit | All Plants | ...
base = OUT
for suffix in ("", "_v2", "_v3", "_v4", "_v5"):
    try:
        OUT = base.replace(".xlsx", f"{suffix}.xlsx")
        wb.save(OUT)
        break
    except PermissionError:  # that version is open in Excel
        continue
with open(os.path.join(SCRATCH, "latest_workbook_path.txt"), "w") as fh:
    fh.write(OUT)
print("saved", OUT)
print("tabs:", wb.sheetnames)
print("screen rows:", len(recs))
